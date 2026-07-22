from __future__ import annotations

import json
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = next(ROOT.glob("*.xlsx"))
OUTPUT = ROOT / "data" / "inventory.json"
REPORT_DATE = date(2026, 7, 22)


def clean(value):
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def normalize(value):
    if value is None:
        return ""
    value = unicodedata.normalize("NFD", str(value).upper())
    value = "".join(char for char in value if unicodedata.category(char) != "Mn")
    return value.replace("Đ", "D").strip()


def classify(reason, condition):
    if condition:
        return "Hàng lỗi / hư"
    reason_normalized = normalize(reason)
    if "THEO DON HANG" in reason_normalized:
        return "Theo đơn hàng"
    if "SX DU" in reason_normalized:
        return "Sản xuất dư"
    return "Chưa xác định"


def age_bucket(days):
    if days <= 7:
        return "0–7 ngày"
    if days <= 30:
        return "8–30 ngày"
    if days <= 60:
        return "31–60 ngày"
    if days <= 90:
        return "61–90 ngày"
    if days <= 180:
        return "91–180 ngày"
    return ">180 ngày"


workbook = load_workbook(SOURCE, data_only=True, read_only=True)
sheet = workbook.active
records = []

for row_number, row in enumerate(sheet.iter_rows(min_row=8, max_row=420, values_only=True), start=8):
    if not row[0]:
        continue
    received = row[16]
    age_days = (REPORT_DATE - received.date()).days if isinstance(received, datetime) else None
    status = classify(row[20], row[21])
    records.append(
        {
            "row": row_number,
            "sku": row[0],
            "product": row[1],
            "barcode": str(row[2]) if row[2] is not None else "",
            "dimensions": {"length": row[4] or 0, "width": row[5] or 0, "height": row[6] or 0},
            "qtyIn": row[7] or 0,
            "qtyOut": row[8] or 0,
            "qtyStock": row[9] or 0,
            "volumeIn": row[10] or 0,
            "volumeOut": row[11] or 0,
            "volumeStock": row[12] or 0,
            "value": row[13] or 0,
            "warehouse": row[14] or "Chưa xác định",
            "note": row[15] or "",
            "receivedAt": clean(received),
            "receiptNo": row[17] or "",
            "issuedAt": clean(row[18]),
            "issueNo": row[19] or "",
            "reason": row[20] or "Chưa xác định",
            "condition": row[21] or "Bình thường",
            "grade": row[22] if row[22] is not None else "Chưa xác định",
            "ageHours": row[23] or 0,
            "warehouseName": row[24] or "",
            "status": status,
            "ageDays": age_days,
            "ageBucket": age_bucket(age_days) if age_days is not None else "Thiếu ngày nhập",
        }
    )

summary = {
    "rows": len(records),
    "qtyStock": sum(item["qtyStock"] for item in records),
    "volumeStock": sum(item["volumeStock"] for item in records),
    "value": sum(item["value"] for item in records),
    "skuCount": len({item["sku"] for item in records}),
    "barcodeCount": len({item["barcode"] for item in records if item["barcode"]}),
}

payload = {
    "meta": {
        "source": SOURCE.name,
        "sheet": sheet.title,
        "reportDate": REPORT_DATE.isoformat(),
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "statusRule": "Tình trạng lỗi được ưu tiên; sau đó phân loại theo lý do nhập.",
    },
    "summary": summary,
    "records": records,
}

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
print(f"Wrote {len(records)} records to {OUTPUT.relative_to(ROOT)}")
